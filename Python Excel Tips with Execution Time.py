#!/usr/bin/env python
# coding: utf-8

# # TOP 5 Ways to Read Excel Data in Python

# In[1]:


excel_path = r"C:\Users\Home\Google Drive\Youtube\1. 5 ways to read excel data using python\Files and COde/Sales.xlsx"


# ### WAY 1: Read excel file with openpyxl

# In[2]:


import openpyxl
wb = openpyxl.load_workbook(excel_path) 

#Get sheet names
print (wb.sheetnames)

#Get rows and column counts from each sheet
for i in wb.sheetnames:
    print("rows in", i , wb[i].max_row)
    print("columns in", i , wb[i].max_column)


# ### WAY 2: Read excel file with xlrd

# In[3]:


import xlrd 
wbk = xlrd.open_workbook(excel_path)

wbk.sheet_names()

for i in wbk.sheet_names():
    print("rows in", i , wbk.sheet_by_name(i).nrows)
    print("columns in", i , wbk.sheet_by_name(i).ncols)


# ### WAY 3: Read excel file with pylightxl

# In[4]:


#!pip install pylightxl

import pylightxl as xl
xlbk = xl.readxl(fn=excel_path)

xlbk.ws_names

for i in xlbk.ws_names:
    print("rows in", i , len(xlbk.ws(ws=i).col(col=1)))
    print("columns in", i , len(xlbk.ws(ws=i).row(row=1)))


# ### WAY 4: Read whole excel file as an object using pandas

# In[5]:



import pandas as pd
xlwb = pd.ExcelFile(excel_path)

xlwb.sheet_names  # see all sheet names

df1 = pd.read_excel(xlwb,sheet_name='Orders')
print(df1.shape)

df2 = pd.read_excel(xlwb,sheet_name='People')
print(df2.shape)

df3 = pd.read_excel(xlwb,sheet_name='Returns')
print(df3.shape)


# ### WAY 5: Read individual sheets without creating objects in pandas

# In[6]:


import pandas as pd

df1 = pd.read_excel(excel_path,sheet_name='Orders')
print(df1.shape)

df2 = pd.read_excel(excel_path,sheet_name='People')
print(df2.shape)

df3 = pd.read_excel(excel_path,sheet_name='Returns')
print(df3.shape)


# In[ ]:




